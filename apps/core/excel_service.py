import openpyxl
from io import BytesIO
from django.http import HttpResponse
from datetime import datetime, date

def export_to_excel(queryset, columns, filename_prefix):
    """
    Export a queryset to an Excel file.
    columns: list of tuples (header_name, field_path)
    field_path can be a simple field name or a related field (e.g., 'hospital.name')
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"

    # Write headers
    for col_num, (header, _) in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data
    for row_num, obj in enumerate(queryset, 2):
        for col_num, (_, field_path) in enumerate(columns, 1):
            value = obj
            for part in field_path.split('.'):
                if value is None:
                    break
                # Handle choices
                if hasattr(value, f'get_{part}_display'):
                    value = getattr(value, f'get_{part}_display')()
                else:
                    value = getattr(value, part, None)
            
            # Format dates
            if isinstance(value, (datetime, date)):
                value = value.strftime('%Y-%m-%d')
            
            ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def parse_excel_file(file_obj, column_map):
    """
    Parse an Excel file and return a list of dictionaries.
    column_map: dict of {header_name: internal_field_name}
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    
    rows = list(ws.rows)
    if not rows:
        return []

    headers = [cell.value for cell in rows[0]]
    header_indices = {header: i for i, header in enumerate(headers) if header in column_map}

    data = []
    for row in rows[1:]:
        row_data = {}
        for header, internal_name in column_map.items():
            if header in header_indices:
                idx = header_indices[header]
                val = row[idx].value
                # Basic cleaning
                if isinstance(val, str):
                    val = val.strip()
                row_data[internal_name] = val
        if any(row_data.values()): # Skip empty rows
            data.append(row_data)
            
    return data
