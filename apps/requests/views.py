from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.shortcuts import redirect
from django.core.exceptions import PermissionDenied
from django.db.models import Q

from .models import Request
from .forms import RequestForm, RequestStatusForm
from apps.users.models import User
from apps.core.models import Status


# 🔒 Статусы, при которых обычный User не может редактировать
CLOSED_STATUSES = [
    'Выполнено',
    'Закрыто',
    'Отменено',
    'Completed',
    'Closed',
]


class RequestListView(LoginRequiredMixin, ListView):
    model = Request
    template_name = 'requests/list.html'
    context_object_name = 'requests'
    ordering = ['-created_at']

    def get_queryset(self):
        queryset = super().get_queryset()

        # 🔒 Filtering for regular users
        if self.request.user.role == User.Role.USER:
            queryset = queryset.filter(responsible=self.request.user).exclude(status__name__in=CLOSED_STATUSES)

        q = self.request.GET.get('q')
        status_id = self.request.GET.get('status')

        if q:
            queryset = queryset.filter(
                Q(request_number__icontains=q) |
                Q(description__icontains=q) |
                Q(responsible__full_name__icontains=q)
            )

        if status_id and status_id.isdigit():
            queryset = queryset.filter(status_id=int(status_id))

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        status = self.request.GET.get('status')
        context['selected_status'] = int(status) if status and status.isdigit() else None

        context['statuses'] = Status.objects.all()
        context['closed_statuses'] = CLOSED_STATUSES
        return context


class RequestDetailView(LoginRequiredMixin, DetailView):
    model = Request
    template_name = 'requests/detail.html'
    context_object_name = 'request_obj'


class RequestCreateView(LoginRequiredMixin, CreateView):
    model = Request
    form_class = RequestForm
    template_name = 'requests/form.html'
    success_url = reverse_lazy('requests:list')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role == User.Role.USER:
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.responsible = self.request.user
        return super().form_valid(form)


class RequestUpdateView(LoginRequiredMixin, UpdateView):
    model = Request
    template_name = 'requests/form.html'
    success_url = reverse_lazy('requests:list')

    def dispatch(self, request, *args, **kwargs):
        response = super().dispatch(request, *args, **kwargs)
        obj = self.object

        if request.user.role == User.Role.USER:
            if obj.status and obj.status.name in CLOSED_STATUSES:
                return redirect('requests:list')

        return response

    def get_form_class(self):
        if self.request.user.role == User.Role.USER:
            return RequestStatusForm
        return RequestForm


import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from apps.core.models import Region, Institution, EquipmentType

def export_requests_to_excel(request):
    """
    Export all requests to an Excel file.
    """
    if request.user.role == User.Role.USER:
        raise PermissionDenied

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=requests_export.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Requests'

    # Header row
    columns = [
        'Request Number', 'Created At', 'Region', 'Institution',
        'Equipment Type', 'Status', 'Responsible', 'Contact Phone', 'Description'
    ]
    worksheet.append(columns)

    for req in Request.objects.all().order_by('-created_at'):
        row = [
            req.request_number,
            req.created_at,
            req.region.name if req.region else '',
            req.institution.name if req.institution else '',
            req.equipment_type.name if req.equipment_type else '',
            req.status.name if req.status else '',
            req.responsible.full_name if req.responsible else '',
            req.contact_phone,
            req.description
        ]
        worksheet.append(row)

    workbook.save(response)
    return response


def import_requests_from_excel(request):
    """
    Import requests from an Excel file.
    """
    if request.user.role == User.Role.USER:
        raise PermissionDenied

    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb.active

        # Assuming header is the first row, start from the second
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # row structure matches export:
            # 0: Request Number (ignored/generated), 1: Created At (ignored/auto), 
            # 2: Region, 3: Institution, 4: Equipment, 5: Status, 
            # 6: Responsible (ignored/current user), 7: Phone, 8: Description
            
            # Simple logic: Lookup FKs by name. If not found, skip or create (here we skip if not found for safety)
            region_name = row[2]
            institution_name = row[3]
            equipment_name = row[4]
            status_name = row[5]
            phone = row[7]
            desc = row[8]

            region = Region.objects.filter(name__iexact=region_name).first()
            institution = Institution.objects.filter(name__iexact=institution_name).first()
            equipment = EquipmentType.objects.filter(name__iexact=equipment_name).first()
            status = Status.objects.filter(name__iexact=status_name).first()

            if region and institution and equipment and status:
                Request.objects.create(
                    region=region,
                    institution=institution,
                    equipment_type=equipment,
                    status=status,
                    contact_phone=phone or '',
                    description=desc or '',
                    responsible=request.user 
                )
        
        return redirect('requests:list')

    return render(request, 'common/import_form.html')
